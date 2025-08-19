#!/usr/bin/env python
# -*- coding: utf-8 -*-
from datetime import datetime
from io import BytesIO

import numpy as np
import pandas
from django.http import HttpResponse
from django.utils.encoding import escape_uri_path
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from xadmin.util import create_event_log

from utils.throttle_cache_tools import get_ip_address


def get_admin_excel_response(
    request,
    opts,
    filename: str,
    data_frame: pandas.DataFrame,
    title: str = None,
    auto_filter: bool = False,
    auto_adaptive_column_width: bool = False,
    sheet_name: str = "Sheet1",
):
    """
    抽取一个公用的pandas DataFrame 导出 excel file  写入 响应头数据中返回

    :param request:                      xadmin request
    :param opts:                        opts
    :param filename:                    文件名称
    :param title:                       文件首行自定义标题
    :param data_frame:                  写入excel内的DataFrame内容数据对象
    :param auto_filter:                 自动为每个字段添加过滤器
    :param auto_adaptive_column_width:  自动为每个字段设置最佳列宽
    :param sheet_name:                  默认的sheet_name
    :return: HttpResponse
    """
    outfile = BytesIO()
    response = HttpResponse(content_type="application/vnd.ms-excel")
    response["Content-Disposition"] = "attachment;filename={0}.xlsx".format(
        escape_uri_path(filename)
    )
    data_length = data_frame.shape[0]
    with pandas.ExcelWriter(outfile, engine="openpyxl") as writer:
        start_row = 0 if title is None else 1
        data_frame.to_excel(
            excel_writer=writer,
            sheet_name=sheet_name,
            index=False,
            freeze_panes=(1, 0),
            startrow=start_row,
        )

        sheet_switch = (
            auto_filter or auto_adaptive_column_width or start_row == 1
        )
        if sheet_switch:
            # 获取底层的Workbook对象
            workbook = writer.book
            # 获取Sheet对象
            sheet = workbook[sheet_name]

        if not data_frame.empty and start_row == 1:
            # 设置第一行为独占的标题行（冻结窗格）
            # sheet.freeze_panes = 'A2'  # 冻结从A2单元格开始的位置
            sheet.freeze_panes = sheet["A3"]  # 冻结第1、2行

            # 插入标题
            title_range = f"A1:{get_column_letter(len(data_frame.columns))}1"
            sheet.merge_cells(title_range)
            sheet[title_range][0][0].value = title
            sheet[title_range][0][0].alignment = Alignment(
                horizontal="center", vertical="center"
            )

            # 设置标题行的行高和字体大小
            title_row = 1
            sheet.row_dimensions[title_row].height = 25  # 设置行高为25
            for cell in sheet[title_row]:
                cell.font = Font(size=14)  # 设置字体大小为14

        if not data_frame.empty and auto_filter:
            # 添加自动筛选器
            # sheet.auto_filter.ref = sheet.dimensions # 所有字段第一行开启自动筛选
            # 第二行字段开启自动筛选
            sheet.auto_filter.ref = (
                f"A2:{get_column_letter(len(data_frame.columns))}2"
            )

        if auto_adaptive_column_width:
            # --以下为了设置自适应列宽
            # 计算表头的字符宽度
            column_widths = (
                data_frame.columns.to_series()
                .apply(lambda x: len(x.encode("utf8")))
                .values
            )

            # 计算每列的最大字符宽度
            max_widths = (
                data_frame.astype(str)
                .map(lambda x: len(x.encode("utf8")))
                .agg("max")
                .values
            )

            # 计算整体最大宽度
            widths = np.max([column_widths, max_widths], axis=0)

            for i, width in enumerate(widths, 1):
                sheet.column_dimensions[get_column_letter(i)].width = width + 3

        # if sheet_switch:
        #     # 保存Excel文件
        #     writer.close()

    response.write(outfile.getvalue())
    ip_addr = get_ip_address(request)
    create_event_log(
        user=request.user,
        flag="export",
        message="导出 %s %s格式 %s条数据" % (filename, "xlsx", data_length),
        ip_addr=ip_addr,
        app_label=opts.app_label,
        model_name=opts.model_name,
        object_repr="数据导出记录",
    )
    return response


def get_admin_excel_response_overtime(
    request,
    opts,
    filename: str,
    data_frame: pandas.DataFrame,
    title: str = None,
    auto_filter: bool = False,
    auto_adaptive_column_width: bool = False,
    sheet_name: str = "Sheet1",
):
    """
    抽取一个公用的pandas DataFrame 导出 excel file  写入 响应头数据中返回

    :param request:                      xadmin request
    :param opts:                        opts
    :param filename:                    文件名称
    :param title:                       文件首行自定义标题
    :param data_frame:                  写入excel内的DataFrame内容数据对象
    :param auto_filter:                 自动为每个字段添加过滤器
    :param auto_adaptive_column_width:  自动为每个字段设置最佳列宽
    :param sheet_name:                  默认的sheet_name
    :return: HttpResponse
    """
    outfile = BytesIO()
    response = HttpResponse(content_type="application/vnd.ms-excel")
    response["Content-Disposition"] = "attachment;filename={0}.xlsx".format(
        escape_uri_path(filename)
    )
    data_length = data_frame.shape[0]
    with pandas.ExcelWriter(outfile, engine="openpyxl") as writer:
        start_row = 0 if title is None else 1
        data_frame.to_excel(
            excel_writer=writer,
            sheet_name=sheet_name,
            index=False,
            freeze_panes=(1, 0),
            startrow=start_row,
        )

        sheet_switch = (
            auto_filter or auto_adaptive_column_width or start_row == 1
        )
        if sheet_switch:
            # 获取底层的Workbook对象
            workbook = writer.book
            # 获取Sheet对象
            sheet = workbook[sheet_name]

        if not data_frame.empty and start_row == 1:
            # 设置第一行为独占的标题行（冻结窗格）
            # sheet.freeze_panes = 'A2'  # 冻结从A2单元格开始的位置
            sheet.freeze_panes = sheet["A3"]  # 冻结第1、2行

            # 插入标题
            title_range = f"A1:{get_column_letter(len(data_frame.columns))}1"
            sheet.merge_cells(title_range)
            sheet[title_range][0][0].value = title
            sheet[title_range][0][0].alignment = Alignment(
                horizontal="center", vertical="center"
            )

            # 设置标题行的行高和字体大小
            title_row = 1
            sheet.row_dimensions[title_row].height = 25  # 设置行高为25
            for cell in sheet[title_row]:
                cell.font = Font(size=14)  # 设置字体大小为14

        if not data_frame.empty and auto_filter:
            # 添加自动筛选器
            # sheet.auto_filter.ref = sheet.dimensions # 所有字段第一行开启自动筛选
            # 第二行字段开启自动筛选
            sheet.auto_filter.ref = (
                f"A2:{get_column_letter(len(data_frame.columns))}2"
            )

        if auto_adaptive_column_width:
            # --以下为了设置自适应列宽
            # 计算表头的字符宽度
            column_widths = (
                data_frame.columns.to_series()
                .apply(lambda x: len(x.encode("utf8")))
                .values
            )

            # 计算每列的最大字符宽度
            max_widths = (
                data_frame.astype(str)
                .map(lambda x: len(x.encode("utf8")))
                .agg("max")
                .values
            )

            # 计算整体最大宽度
            widths = np.max([column_widths, max_widths], axis=0)

            for i, width in enumerate(widths, 1):
                sheet.column_dimensions[get_column_letter(i)].width = width + 3

        lenstr=len(data_frame)
        if not data_frame.empty and len(data_frame)>0:
            # 获取底层的Workbook对象
            workbook = writer.book
            # 获取Sheet对象
            sheet = workbook[sheet_name]
            # 找到 "配件照片" 列的索引
            att_column_index = data_frame.columns.get_loc("配件照片") + 1  # +1 使索引从 1 开始
            att_column_letter = get_column_letter(att_column_index)  # 获取列字母
            for index, item in enumerate(data_frame.itertuples(index=False), start=2):
                # 确保通过 item 访问到 attachment 字段
                attachment_url = getattr(item, "配件照片")  # 获取"配件照片"字段的值
                # 仅当有链接时才设置超链接
                if attachment_url:
                    sheet[f'{att_column_letter}{index}'].hyperlink = attachment_url

        # if sheet_switch:
        #     # 保存Excel文件
        #     writer.close()

    response.write(outfile.getvalue())
    ip_addr = get_ip_address(request)
    create_event_log(
        user=request.user,
        flag="export",
        message="导出 %s %s格式 %s条数据" % (filename, "xlsx", data_length),
        ip_addr=ip_addr,
        app_label=opts.app_label,
        model_name=opts.model_name,
        object_repr="数据导出记录",
    )
    return response

def get_admin_excel_response_overtime_xlsxwriter(
    request,
    opts,
    filename: str,
    data_frame: pandas.DataFrame,
    title: str = None,
    auto_filter: bool = False,
    auto_adaptive_column_width: bool = False,
    sheet_name: str = "Sheet1",
):
    """
    抽取一个公用的pandas DataFrame 导出 excel file  写入 响应头数据中返回

    :param request:                      xadmin request
    :param opts:                        opts
    :param filename:                    文件名称
    :param title:                       文件首行自定义标题
    :param data_frame:                  写入excel内的DataFrame内容数据对象
    :param auto_filter:                 自动为每个字段添加过滤器
    :param auto_adaptive_column_width:  自动为每个字段设置最佳列宽
    :param sheet_name:                  默认的sheet_name
    :return: HttpResponse
    """
    outfile = BytesIO()
    response = HttpResponse(content_type="application/vnd.ms-excel")
    response["Content-Disposition"] = "attachment;filename={0}.xlsx".format(
        escape_uri_path(filename)
    )
    data_length = data_frame.shape[0]
    with pandas.ExcelWriter(outfile, engine="xlsxwriter") as writer:
        start_row = 0 if title is None else 1
        data_frame.to_excel(
            excel_writer=writer,
            sheet_name=sheet_name,
            index=False,
            freeze_panes=(1, 0),
            startrow=start_row,
        )
        lenstr=len(data_frame)
        if not data_frame.empty and len(data_frame)>0:
            # 获取底层的Workbook对象
            workbook = writer.book
            # 获取Sheet对象
            sheet = workbook.get_worksheet_by_name(sheet_name)
            # 找到 "配件照片" 列的索引
            att_column_index = data_frame.columns.get_loc("配件照片") + 1  # +1 使索引从 1 开始
            att_column_letter = get_column_letter(att_column_index)  # 获取列字母
            for index, item in enumerate(data_frame.itertuples(index=False), start=2):
                # 确保通过 item 访问到 attachment 字段
                attachment_url = getattr(item, "配件照片")  # 获取"配件照片"字段的值
                # 仅当有链接时才设置超链接
                if attachment_url:
                    #sheet[f'{att_column_letter}{index}'].hyperlink = attachment_url
                    sheet.write_url(f'{att_column_letter}{index}', attachment_url)

    response.write(outfile.getvalue())
    ip_addr = get_ip_address(request)
    create_event_log(
        user=request.user,
        flag="export",
        message="导出 %s %s格式 %s条数据" % (filename, "xlsx", data_length),
        ip_addr=ip_addr,
        app_label=opts.app_label,
        model_name=opts.model_name,
        object_repr="数据导出记录",
    )
    return response


